import os, sys
from openpyxl import Workbook,load_workbook


wb = Workbook()
wb = load_workbook('a.xlsx')
w_s=wb[wb.sheetnames[0]]
rows=w_s.iter_rows(min_row=2,min_col=2,max_col=2)
l=[]
i=0
for row in rows:
    for cell in row:
        b = os.popen('adb shell dumpsys package %s | findstr "versionName"' %cell.value).read()
        c=b.split('=')[1].replace('\n','')
        print(c)
        l.append(c)
print("totoal number= "+str(len(l)))
rows3=w_s.iter_rows(min_row=2,min_col=3,max_col=3)
for row in rows3:
    for cell in row:
        # print(i)
        cell.value=l[i]
        i=i+1
wb.save('a.xlsx')